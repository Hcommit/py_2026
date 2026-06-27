""""

Consider a studData.xlsx file. The file has the USN, Name and CGPA of the students in the class. 
Develop a program to find the first topper of the class.
"""




from openpyxl import load_workbook

# Load the Excel file
wb = load_workbook(r"C:\Users\Hardik Gupta\OneDrive\Desktop\CODING\ZZZZZZZZZZZ LAB INTERNAL\pythonLABS\studyData.xlsx")
ws = wb.active

# Initialize variables to track topper
topperUsn = ""
topperName = ""
topperCgpa = -1

# Loop through the rows (assuming header in first row)
for row in ws.iter_rows(min_row=2, values_only=True):
    usn, name, cgpa = row

    if cgpa > topperCgpa:
        topperCgpa = cgpa
        topperName = name
        topperUsn = usn

# Display first topper
print("--- First Topper of the Class ---")
print("USN:", topperUsn)
print("Name:", topperName)
print("CGPA:", topperCgpa)